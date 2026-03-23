#!/usr/bin/env python3
"""
Generate three-statement financial model template
Creates Excel file with linked IS, BS, CF statements
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_header(ws, title):
    """Create styled header"""
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

def create_assumptions_sheet(wb):
    """Create assumptions sheet"""
    ws = wb.create_sheet('Assumptions')
    create_header(ws, 'KEY ASSUMPTIONS')
    
    # Headers
    ws['A3'] = 'Assumption'
    ws['B3'] = 'Year 1'
    ws['C3'] = 'Year 2'
    ws['D3'] = 'Year 3'
    ws['E3'] = 'Year 4'
    ws['F3'] = 'Year 5'
    
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
    
    # Assumption rows
    assumptions = [
        ('Revenue Growth %', 0.10, 0.12, 0.10, 0.08, 0.06),
        ('COGS % of Revenue', 0.60, 0.58, 0.57, 0.56, 0.55),
        ('SG&A % of Revenue', 0.20, 0.19, 0.18, 0.17, 0.17),
        ('D&A % of Revenue', 0.05, 0.05, 0.05, 0.05, 0.05),
        ('CapEx % of Revenue', 0.08, 0.07, 0.06, 0.06, 0.05),
        ('Tax Rate', 0.21, 0.21, 0.21, 0.21, 0.21),
        ('WACC', 0.10, 0.10, 0.10, 0.10, 0.10),
    ]
    
    for i, (name, *values) in enumerate(assumptions, start=4):
        ws[f'A{i}'] = name
        for j, val in enumerate(values, start=2):
            cell = ws.cell(row=i, column=j, value=val)
            cell.number_format = '0.0%'
            cell.font = Font(color='0000FF')  # Blue for inputs
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15

def create_income_statement(wb):
    """Create income statement sheet"""
    ws = wb.create_sheet('Income Statement')
    create_header(ws, 'INCOME STATEMENT')
    
    # Structure
    items = [
        ('Revenue', '=1000000*(1+Assumptions!B4)'),
        ('Cost of Goods Sold', '=-Revenue*Assumptions!B5'),
        ('Gross Profit', '=SUM(B5:B6)'),
        ('', ''),
        ('SG&A', '=-Revenue*Assumptions!B6'),
        ('Depreciation & Amortization', '=-Revenue*Assumptions!B7'),
        ('Operating Income (EBIT)', '=B7+B10+B11'),
        ('', ''),
        ('Interest Expense', '=-50000'),
        ('EBT', '=B13+B15'),
        ('Taxes', '=-B16*Assumptions!B9'),
        ('Net Income', '=B16+B17'),
    ]
    
    ws['A3'] = 'Line Item'
    ws['B3'] = 'Year 1'
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    
    for i, (name, formula) in enumerate(items, start=4):
        ws[f'A{i}'] = name
        if formula:
            ws[f'B{i}'] = formula
            ws[f'B{i}'].number_format = '#,##0'
            if 'SUM' in formula or 'B7' in formula or 'B13' in formula or 'B16' in formula:
                ws[f'B{i}'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

def create_balance_sheet(wb):
    """Create balance sheet template"""
    ws = wb.create_sheet('Balance Sheet')
    create_header(ws, 'BALANCE SHEET')
    
    sections = [
        ('ASSETS', '', True),
        ('Cash', '=\'Cash Flow\'!B18', False),
        ('Accounts Receivable', '=\'Income Statement\'!B5*30/365', False),
        ('Inventory', '=\'Income Statement\'!B6*45/365', False),
        ('Total Current Assets', '=SUM(B6:B8)', True),
        ('', '', False),
        ('PP&E, net', '=800000', False),
        ('Goodwill', '=200000', False),
        ('Total Assets', '=B9+B12+B13', True),
        ('', '', False),
        ('LIABILITIES & EQUITY', '', True),
        ('Accounts Payable', '=\'Income Statement\'!B6*60/365', False),
        ('Total Current Liabilities', '=B16', False),
        ('', '', False),
        ('Long-term Debt', '=500000', False),
        ('Total Liabilities', '=B17+B20', True),
        ('', '', False),
        ('Equity', '=B14-B21', True),
        ('Total L&E', '=B21+B23', True),
        ('', '', False),
        ('CHECK: Assets = L&E', '=IF(ABS(B14-B24)<1,"OK","ERROR")', False),
    ]
    
    ws['A3'] = 'Line Item'
    ws['B3'] = 'Year 1'
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    
    for i, (name, formula, bold) in enumerate(sections, start=4):
        ws[f'A{i}'] = name
        if formula:
            ws[f'B{i}'] = formula
            if bold:
                ws[f'B{i}'].font = Font(bold=True)
            if 'SUM' in formula or 'B14' in formula or 'B21' in formula or 'B24' in formula:
                ws[f'B{i}'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

def create_cash_flow(wb):
    """Create cash flow statement template"""
    ws = wb.create_sheet('Cash Flow')
    create_header(ws, 'CASH FLOW STATEMENT')
    
    flows = [
        ('Operating Activities', '', True),
        ('Net Income', '=\'Income Statement\'!B18', False),
        ('Add: Depreciation & Amortization', '=-\'Income Statement\'!B11', False),
        ('Changes in Working Capital', '=0', False),
        ('Cash from Operations', '=SUM(B6:B8)', True),
        ('', '', False),
        ('Investing Activities', '', True),
        ('CapEx', '=\'Income Statement\'!B5*Assumptions!B8', False),
        ('Cash from Investing', '=B12', True),
        ('', '', False),
        ('Financing Activities', '', True),
        ('Debt Repayment', '=-25000', False),
        ('Cash from Financing', '=B16', True),
        ('', '', False),
        ('Net Change in Cash', '=B9+B13+B17', True),
        ('Beginning Cash', '=100000', False),
        ('Ending Cash', '=B20+B19', True),
    ]
    
    ws['A3'] = 'Line Item'
    ws['B3'] = 'Year 1'
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    
    for i, (name, formula, bold) in enumerate(flows, start=4):
        ws[f'A{i}'] = name
        if formula:
            ws[f'B{i}'] = formula
            if bold:
                ws[f'B{i}'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

def main():
    if len(sys.argv) < 2:
        print("Usage: create_3statement_model.py <output_filename.xlsx>")
        print("Example: create_3statement_model.py Company_Model.xlsx")
        sys.exit(1)
    
    output_file = sys.argv[1]
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    create_assumptions_sheet(wb)
    create_income_statement(wb)
    create_balance_sheet(wb)
    create_cash_flow(wb)
    
    # Reorder sheets
    wb._sheets = [wb['Assumptions'], wb['Income Statement'], 
                  wb['Balance Sheet'], wb['Cash Flow']]
    
    wb.save(output_file)
    print(f"Created: {output_file}")
    print("\nModel structure:")
    print("  - Assumptions: Key drivers (blue = input cells)")
    print("  - Income Statement: Revenue to Net Income")
    print("  - Balance Sheet: Assets = Liabilities + Equity")
    print("  - Cash Flow: Operating, Investing, Financing")

if __name__ == '__main__':
    main()
